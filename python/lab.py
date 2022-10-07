import os
from re import sub
# import pandas as pd
from openpyxl import load_workbook

from openpyxl.styles import Alignment
from openpyxl.styles.fonts import Font

path = "../python/raw"

file_list = os.listdir(path)

title_list = ['UPC', 'BRAND', 'DESC1', 'DESC2', 'SIZE', 'FAMILY', 'FNAME', 'REPORT', 'RNAME', 'SDEPT', 'SNAME',
 'CATEGORY', 'CNAME', 'TAX1' ,'TAX2', 'FSTMP', 'SCALE', 'MANUAL', 'PLU', 'VENDOR', 'VNAME', 'AUTH', 'VENDOR CODE',
  'BASECOST', ' CASEQTY', 'UNITCOST', 'ACTIVE', 'REG', 'TPR', 'SALE', 'NFS']


for f in file_list:

  file_name = "../python/raw/" + f
  wb = load_workbook(filename = file_name, data_only=True)
  ws = wb.active

  last_raw = ws.max_row
 

  result = {}

  for i in range(2, last_raw):
      if result.get(ws['A'+str(i)].value) == None:
          result[ws['A'+str(i)].value] = []
      result[ws['A'+str(i)].value].append([str(ws['A'+str(i)].value), ws['B'+str(i)].value, ws['C'+str(i)].value, ws['D'+str(i)].value, ws['E'+str(i)].value, ws['F'+str(i)].value, ws['G'+str(i)].value,
       ws['H'+str(i)].value, ws['I'+str(i)].value, ws['J'+str(i)].value, ws['K'+str(i)].value, ws['L'+str(i)].value, ws['M'+str(i)].value, ws['N'+str(i)].value, ws['O'+str(i)].value, ws['P'+str(i)].value,
        ws['Q'+str(i)].value, ws['R'+str(i)].value, ws['S'+str(i)].value, ws['T'+str(i)].value, ws['U'+str(i)].value, ws['V'+str(i)].value, ws['W'+str(i)].value, float(ws['X'+str(i)].value),
        "{:.2f}".format((float(ws['Y'+str(i)].value))), "{:.2f}".format((float(ws['Z'+str(i)].value))), "{:.2f}".format((float(ws['AA'+str(i)].value))),
         "{:.2f}".format((float(ws['AB'+str(i)].value))),"{:.2f}".format((float(ws['AC'+str(i)].value))),"{:.2f}".format((float(ws['AD'+str(i)].value))), ws['AE'+str(i)].value])

#   print(result)
  from openpyxl import Workbook
  wb = Workbook()


  temp_list= []
  for i in result.keys():
    for j in range(0, len(result[i])):
        if str(result[i][j][30]) == '1' or \
          str(result[i][j][1]).lower() in ['happy hour', 'coupon', 'event'] or \
          str(result[i][j][0]) in ['0000000000091', '0000000000092', '0000000000093', '0000000000094','0000000000095', '0000000000096', '0000000000097', '0000000000098', '0000000000999', '0000000009999', '0000000099999'] or \
          str(result[i][j][2]) in ['ASSI SERVICE CHARGE', 'BOTTLE REFUND ITEM', 'H/W SERVICE FEE', 'BEER TEST', 'WINE TEST'] or \
          str(result[i][j][7]) in ['99', '999'] or result[i][j][7]==None :
            temp_list.append(result[i][j][0])
  for i in temp_list:
    result.pop(i, None)


  #MRAGIN

  if 'GA55' in f:
    ws0 = wb.create_sheet("MARGIN")
    ga55_margin = []

  margin_title_list = ['UPC', 'BRAND', 'DESC1', 'DESC2', 'SIZE', 'FAMILY', 'FNAME', 'REPORT', 'RNAME', 'SDEPT', 'SNAME',
  'CATEGORY', 'CNAME', 'TAX1' ,'TAX2', 'FSTMP', 'SCALE', 'MANUAL', 'PLU', 'VENDOR', 'VNAME', 'AUTH', 'VENDOR CODE',
    'BASECOST', ' CASEQTY', 'UNITCOST', 'ACTIVE', 'REG', 'TPR', 'SALE', 'MARGIN']

  for i in range(1, 32):
      ws0.cell(row=1, column=i).value = margin_title_list[i-1]
      ws0.cell(row=1, column=i).alignment = Alignment(horizontal = 'center', vertical = 'center')
      ws0.cell(row=1, column=i).font = Font(bold=True)


  for i in result.keys():
    for j in range(0, len(result[i])):
        if float(result[i][j][26]) != 0:
          if ((float(result[i][j][26])-float(result[i][j][25]))/float(result[i][j][26]))*100 <=20 and result[i][j][7] in ['11', '12', '14', '17'] and result[i][j][21] == '1':
            ga55_margin.append(result[i][j])
 
  margin_temp = len(ga55_margin)   
  for i in range(1, margin_temp+1):
      for j in range(1, 24):
          ws0.cell(row=i+1, column=j).value = ga55_margin[i-1][j-1]  
      for j in range(24, 31):
          ws0.cell(row=i+1, column=j).value = "{:.2f}".format(float(ga55_margin[i-1][j-1]))  
      ws0.cell(row=i+1, column=31).value = "{:.2f}".format(((float(ga55_margin[i-1][26])-float(ga55_margin[i-1][25]))/float(ga55_margin[i-1][26])) * 100)
    

  #Auth
  ws = wb.active
  subList1 = [] # UPC 한개, Auth 코드 1 (정상 데이터)
  subList2 = [] # UPC 2개 이상인 전체 데이터
  subList3 = [] # UPC 한개, Auth 코드 0 or NULL(blank)(오류)

  subList4 = [] #UPC 2개 이상, Auth 코드 모두 0 (오류)
  subList5 = [] #UPC 2개 이상, VNAME이 다른 AUth 코드 1이 2개 이상 (오류)
  subList6 = [] #UPC 2개 이상, Auto code에 NULL(blank)값이 한개라도 있는 경우 (오류)



  

  for i in result.keys():
    
      #UPC 한 개에, Auth 코드 1인 경우 -> 정상 데이터
      
      if len(result[i])==1 :
          if result[i][0][21]=='1':
              subList1 = result[i] + subList1
            #UPC가 한 개인데, Auth 코드가 0 or NULL인 경우 -> 오류
          else :
              subList3 = result[i] + subList3
      # #UPC가 두 개이상인 경우
      else :
          subList2 = result[i] + subList2
          tempList = []
          tempList2 = []
          for j in range(0, len(result[i])):
              tempList.append(result[i][j][21])
              tempList2.append([result[i][j][20], result[i][j][21]]) #VNAME, Auth code
          
          #UPC가 여러개인데, Auth code가 모두 0인 경우 -> 오류
          vendorList=[]
          if  tempList.count('1')==0:
              subList4 = result[i] + subList4
       
          #Auth code에 NULL값이 있는 데이터의 경우 모두 가져오기 -> 오류
          elif tempList.count('')!=0:
            subList6 = result[i] + subList6

          #Auth code가 1이 한 개 이상인 경우
          else:
            # vendorList = []
            for k in range(0, len(tempList2)):
              if tempList2[k][1] == '1':
                if vendorList.count(tempList2[k][0]) == 0:
                    vendorList.append(tempList2[k][0])
          #Auth code가 1이 2개 이상인데, vendor name이 다른 경우 - > auth code 오류    
          if len(vendorList) > 1 :
              subList5 = result[i] + subList5

                  
  # 각 subList에 발생한 None값 제거
        

  subList = subList3 + subList4 + subList5 + subList6 #오류 데이터를 모두 합친 리스트

  # 합치기 전에 None값을 제거 하는 속도가 더 빠름
  # while [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None] in subList:
  #      subList.remove([None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None])
  auth_subList = []
  for i in subList:
    if i[0][0:3] != '002':
      auth_subList.append(i)
    elif i[0][0:3] == '002' and 100 <= int(i[0][3:8]) < 5000 :
      auth_subList.append(i)

  # 로젠 국제 특송 제거
  auth_subList = [item for item in auth_subList if item[2] != 'LOGEN AMERICA EXPRESS']
  # 상품권 UPC제거
  auth_subList = [item for item in auth_subList if item[0] not in ['0419938191009', '0419938191054', '0419938192006', '0419938192051', '0419938195007', '0419938195052']]
  if 'IL70' in f:
    auth_subList = [item for item in auth_subList if item[0] not in ['0007330920007','0007330920020']]  
  #UPC기준으로 오름차순 정렬
  subList.sort(key=lambda x:x[0])
  auth_subList.sort(key=lambda x:x[0])
  temp = len(auth_subList)
  #print(subList)

  #제목 입력
  auth_title_list = ['UPC', 'BRAND', 'DESC1', 'DESC2', 'SIZE', 'VENDOR', 'VNAME', 'AUTH']
  for i in range(1, 9):
      ws.cell(row=1, column=i).value = auth_title_list[i-1]
      ws.cell(row=1, column=i).alignment = Alignment(horizontal = 'center', vertical = 'center')
      ws.cell(row=1, column=i).font = Font(bold=True)

  #새로운 엑셀 파일에 오류 리스트 값  차례대로 입력
  for i in range(1, temp+1):
      for j in range(1, 6):
          ws.cell(row=i+1, column=j).value = auth_subList[i-1][j-1]  
      for j in range(6, 9):
          ws.cell(row=i+1, column=j).value = auth_subList[i-1][j+13]  
  ws.title = "AUTH"     

#   #AUTH2
#   ws0 = wb.create_sheet("AUTH2")
#   auth2_title_list = ['UPC', 'BRAND', 'DESC1', 'DESC2', 'SIZE', 'FAMILY', 'FNAME', 'REPORT', 'RNAME', 'SDEPT', 'SNAME',
#  'CATEGORY', 'CNAME', 'TAX1' ,'TAX2', 'FSTMP', 'SCALE', 'MANUAL', 'PLU', 'VENDOR', 'VNAME', 'AUTH', 'VENDOR CODE',
#   'BASECOST', ' CASEQTY', 'UNITCOST', 'ACTIVE', 'REG', 'TPR', 'SALE']

#   for i in range(1, 31):
#       ws0.cell(row=1, column=i).value = auth2_title_list[i-1]
#       ws0.cell(row=1, column=i).alignment = Alignment(horizontal = 'center', vertical = 'center')
#       ws0.cell(row=1, column=i).font = Font(bold=True)

#   #새로운 엑셀 파일에 오류 리스트 값  차례대로 입력
#   for i in range(1, temp+1):
#       for j in range(1, 31):
#           ws0.cell(row=i+1, column=j).value = auth_subList[i-1][j-1]  

  #BASE COST - 0
  ws1 = wb.create_sheet("BASE COST-0")
  base_title_list = ['UPC', 'BRAND', 'DESC1', 'DESC2', 'SIZE', 'VENDOR', 'VNAME', 'AUTH', 'VENDOR CODE',
  'BASECOST', ' CASEQTY', 'UNITCOST', 'ACTIVE', 'REG', 'TPR', 'SALE']
  
  base_subList1 = []

  if 'GA55' in f :
    for i in result.keys():
      for j in range(0, len(result[i])):
          if result[i][j][23] == 0 and result[i][j][8] not in ['MEAT', 'FISH', 'DELI'] and result[i][j][7] not in ['80', '81', '82', '83', '84', '85', '86', '87', '88', '89'] and result[i][j][7] not in ['30', '40', '50', '60'] and result[i][j][18] == None and result[i][j][20] != 'MISSY COSMETICS & GIFTS INC' and result[i][j][0] not in ['0880803313082', '0885252200975'] :
            base_subList1.append(result[i][j])
  else:
    for i in result.keys():
      for j in range(0, len(result[i])):
          if float(result[i][j][23]) == 0 and result[i][j][8] not in ['MEAT', 'FISH', 'DELI'] and result[i][j][7] not in ['80', '81', '82', '83', '84', '85', '86', '87', '88', '89']:
            base_subList1.append(result[i][j])

  if 'GA55' in f:
    base_subList1 = [item for item in base_subList1 if item[0] not in ['0880777904903', '0880777909565', '0880777909566', '0880777909567', '0880962684035']] 

  if 'PA88' in f:
    base_subList1 = [item for item in base_subList1 if item[0] not in ['0020099900000']] 
 
  if 'IL70' in f:
    base_subList1 = [item for item in base_subList1 if item[0] not in ['0020274000000', '0007330920007', '0007330920020']]    
    base_subList1 = [item for item in base_subList1 if item[0] not in ['0000091202002', '0001428500193', '0001428500256','0001622900058', '0002800031601', '0005210008901', '0005210008911', '0007161590149', '0007980911645', '0008539656443', '0009678501329', '0063626117070', '0073796400020', '0073796400040', '0073796400043', '0073796400063', '0074848510008', '0074848510009', '0074848510013', '0074848510014', '0074848510146', '0074848560106', '0074848560107', '0084255200002', '0084255200084', '0480001621838', '0480001630501', '0480004051117', '0480007010146', '0480011000672', '0480011600401', '0480011604302', '0480011609802', '0480011609902', '0480021012301', '0480021030013', '0480021705101', '0480166860451', '0480168810351', '0480168888010', '0480650618138', '0480650618139', '0480650618141', '0480650618142', '0480650618143', '0480651202080', '0480651374012', '0480651374232', '0480651374233', '0480777027151', '0480888703078']]    
    base_subList1 = [item for item in base_subList1 if item[0] not in ['0880932462659', '0880932462660']]





 
  base_temp = len(base_subList1)   
  base_subList1.sort(key=lambda x:x[0])
  for i in range(1, 17):
      ws1.cell(row=1, column=i).value = base_title_list[i-1]
      ws1.cell(row=1, column=i).alignment = Alignment(horizontal = 'center', vertical = 'center')
      ws1.cell(row=1, column=i).font = Font(bold=True)

  for i in range(1, base_temp+1):
    ws1.cell(row=i+1, column=1).value = base_subList1[i-1][0]
    ws1.cell(row=i+1, column=2).value = base_subList1[i-1][1] 
    ws1.cell(row=i+1, column=3).value = base_subList1[i-1][2] 
    ws1.cell(row=i+1, column=4).value = base_subList1[i-1][3]  
    ws1.cell(row=i+1, column=5).value = base_subList1[i-1][4]
    ws1.cell(row=i+1, column=6).value = base_subList1[i-1][19] 
    ws1.cell(row=i+1, column=7).value = base_subList1[i-1][20]  
    ws1.cell(row=i+1, column=8).value = base_subList1[i-1][21]  
    ws1.cell(row=i+1, column=9).value = base_subList1[i-1][22]  
    ws1.cell(row=i+1, column=10).value = "{:.2f}".format(float(base_subList1[i-1][23]))  
    ws1.cell(row=i+1, column=11).value = base_subList1[i-1][24]  
    ws1.cell(row=i+1, column=12).value = base_subList1[i-1][25]  
    ws1.cell(row=i+1, column=13).value = base_subList1[i-1][26]
    ws1.cell(row=i+1, column=14).value = base_subList1[i-1][27] 
    ws1.cell(row=i+1, column=15).value = base_subList1[i-1][28] 
    ws1.cell(row=i+1, column=16).value = base_subList1[i-1][29]



  #SUB-DEPT
  ws2 = wb.create_sheet("SUB-DEPT")
  sdp_title_list = ['UPC', 'BRAND', 'DESC1', 'DESC2', 'SIZE', 'FAMILY', 'FNAME', 'REPORT', 'RNAME', 'SDEPT', 'SNAME', 'CATEGORY', 'CNAME']
  sdp_subList1 = []
 

  for i in result.keys():
    for j in range(0, len(result[i])):
      if result[i][j][9]=='9999' or result[i][j][9] == None:
         sdp_subList1.append(result[i][j])


  sdp_temp = len(sdp_subList1)
  sdp_subList1.sort(key=lambda x:x[0])

  for i in range(1, 14):
      ws2.cell(row=1, column=i).value = sdp_title_list[i-1]
      ws2.cell(row=1, column=i).alignment = Alignment(horizontal = 'center', vertical = 'center')
      ws2.cell(row=1, column=i).font = Font(bold=True)

  for i in range(1, sdp_temp+1):
    ws2.cell(row=i+1, column=1).value = sdp_subList1[i-1][0]
    ws2.cell(row=i+1, column=2).value = sdp_subList1[i-1][1] 
    ws2.cell(row=i+1, column=3).value = sdp_subList1[i-1][2] 
    ws2.cell(row=i+1, column=4).value = sdp_subList1[i-1][3] 
    ws2.cell(row=i+1, column=5).value = sdp_subList1[i-1][4] 
    ws2.cell(row=i+1, column=6).value = sdp_subList1[i-1][5] 
    ws2.cell(row=i+1, column=7).value = sdp_subList1[i-1][6] 
    ws2.cell(row=i+1, column=8).value = sdp_subList1[i-1][7] 
    ws2.cell(row=i+1, column=9).value = sdp_subList1[i-1][8] 
    ws2.cell(row=i+1, column=10).value = sdp_subList1[i-1][9] 
    ws2.cell(row=i+1, column=11).value = sdp_subList1[i-1][10]
    ws2.cell(row=i+1, column=12).value = sdp_subList1[i-1][11]  
    ws2.cell(row=i+1, column=13).value = sdp_subList1[i-1][12]

          
  
  
  #UPC Check
  ws3 = wb.create_sheet("UPC 체크")
  upc_title_list = ['UPC', 'BRAND', 'DESC1', 'DESC2', 'SIZE', 'REPORT', 'RNAME', 'SDEPT', 'SNAME',
 'CATEGORY', 'CNAME', 'TAX1' ,'TAX2', 'FSTMP', 'SCALE', 'MANUAL', 'PLU', 'VENDOR', 'VNAME', 'AUTH', 'VENDOR CODE',
  'BASECOST', ' CASEQTY', 'UNITCOST', 'ACTIVE', 'REG', 'TPR', 'SALE']

  upc_subList1 = []

  for i in result.keys():
    if result[i][0][0] != None:
      if result[i][0][0] < '0000000999999':
        upc_subList1 += result[i]


  if 'GA55' in f:
    upc_subList1 = [item for item in upc_subList1 if item[0] not in ['0000000041345', '0000000070035']]    
  if 'IL70' in f:
    upc_subList1 = [item for item in upc_subList1 if item[0] not in ['0000000030010', '0000000030011', '0000000500302', '0000000500705', '0000000500909', '0000000500910', '0000000500911', '0000000500912', '0000000500913', '0000000500944', '0000000500945',
    '0000000500946', '0000000501017', '0000000501027', '0000000501028']]
  if 'PA88' in f:
    upc_subList1 = [item for item in upc_subList1 if item[0] not in ['0000000142858', '0000000789110', '0000000789120', '0000000866053' ,'0000000413924', '0000000815536']]    

  upc_temp = len(upc_subList1)
  upc_subList1.sort(key=lambda x:x[0])

  for i in range(1, 22):
      ws3.cell(row=1, column=i).value = upc_title_list[i-1]
      ws3.cell(row=1, column=i).alignment = Alignment(horizontal = 'center', vertical = 'center')
      ws3.cell(row=1, column=i).font = Font(bold=True)

  for i in range(1, upc_temp+1):
      for j in range(1, 6):
          ws3.cell(row=i+1, column=j).value = upc_subList1[i-1][j-1]  

  for i in range(1, upc_temp+1):
      for j in range(6, 22):
          ws3.cell(row=i+1, column=j).value = upc_subList1[i-1][j+1]  

  #PLU
  ws4 = wb.create_sheet("PLU")
  plu_title_list = ['UPC', 'BRAND', 'DESC1', 'DESC2', 'SIZE', 'PLU']

  plu_subList1 = []
 
  for i in result.keys():
    for j in range(0, len(result[i])):
      if result[i][j][18] != None and result[i][j][0][0:3] != '002' :
        plu_subList1.append(result[i][j])

  plu_temp = len(plu_subList1)
  plu_subList1.sort(key=lambda x:x[0])

  for i in range(1, 7):
      ws4.cell(row=1, column=i).value = plu_title_list[i-1]
      ws4.cell(row=1, column=i).alignment = Alignment(horizontal = 'center', vertical = 'center')
      ws4.cell(row=1, column=i).font = Font(bold=True)

  for i in range(1, plu_temp+1):
    ws4.cell(row=i+1, column=1).value = plu_subList1[i-1][0] 
    ws4.cell(row=i+1, column=2).value = plu_subList1[i-1][1]  
    ws4.cell(row=i+1, column=3).value = plu_subList1[i-1][2]  
    ws4.cell(row=i+1, column=4).value = plu_subList1[i-1][3]  
    ws4.cell(row=i+1, column=5).value = plu_subList1[i-1][4]  
    ws4.cell(row=i+1, column=6).value = plu_subList1[i-1][18]   
    
  #SCALE-MANUAL
  ws5 = wb.create_sheet("SCALE-MANUAL")
  scale_title_list = ['UPC', 'BRAND', 'DESC1', 'DESC2', 'SIZE', 'SCALE', 'MANUAL', 'PLU']
  scale_subList1 = []
  
  for i in result.keys():
    for j in range(0, len(result[i])):
      # if result[i][j][16] != result[i][j][17]:
      #   scale_subList1.append(result[i][j])
      if result[i][j][16]=='1' and result[i][j][17]==None:
         scale_subList1.append(result[i][j])
      elif result[i][j][16]=='Y':
         scale_subList1.append(result[i][j])  
      elif result[i][j][16]==None and result[i][j][17]=='1':
         scale_subList1.append(result[i][j])
      elif result[i][j][16]=='N' and result[i][j][17]=='1':
         scale_subList1.append(result[i][j])
  
  if 'GA55' in f:
    scale_subList1 = [item for item in scale_subList1 if item[0] not in ['0020413900000', '0020414000000','0020414100000', '0020415900000', '0020424900000', '0020427000000', '0020432500000', '0020440000000', '0020440100000', '0020440700000', '0020442600000', '0020442800000', '0020455600000', '0020459100000', '0020465100000', '0020470100000', '0020470700000', '0020483900000', '0020485200000', '0020489500000', '0020490800000', '0020493000000', '0020494600000', '0020492800000']]    
  
  scale_temp = len(scale_subList1)
  scale_subList1.sort(key=lambda x:x[0])

  for i in range(1, 9):
      ws5.cell(row=1, column=i).value = scale_title_list[i-1]
      ws5.cell(row=1, column=i).alignment = Alignment(horizontal = 'center', vertical = 'center')
      ws5.cell(row=1, column=i).font = Font(bold=True)

  for i in range(1, scale_temp+1):
    ws5.cell(row=i+1, column=1).value = scale_subList1[i-1][0] 
    ws5.cell(row=i+1, column=2).value = scale_subList1[i-1][1]  
    ws5.cell(row=i+1, column=3).value = scale_subList1[i-1][2]  
    ws5.cell(row=i+1, column=4).value = scale_subList1[i-1][3]  
    ws5.cell(row=i+1, column=5).value = scale_subList1[i-1][4]  
    ws5.cell(row=i+1, column=6).value = scale_subList1[i-1][16]  
    ws5.cell(row=i+1, column=7).value = scale_subList1[i-1][17]  
    ws5.cell(row=i+1, column=8).value = scale_subList1[i-1][18]   


  #DESC2
  ws6 = wb.create_sheet("DESC2")
  desc2_title_list = ['UPC', 'BRAND', 'DESC1', 'DESC2', 'SIZE']
  desc2_subList1 = []
 

  for i in result.keys():
    for j in range(0, len(result[i])):
      if result[i][j][3]==None:
         desc2_subList1.append(result[i][j])

  
  desc2_temp = len(desc2_subList1)
  desc2_subList1.sort(key=lambda x:x[0])

  for i in range(1, 6):
      ws6.cell(row=1, column=i).value = desc2_title_list[i-1]
      ws6.cell(row=1, column=i).alignment = Alignment(horizontal = 'center', vertical = 'center')
      ws6.cell(row=1, column=i).font = Font(bold=True)

  for i in range(1, desc2_temp+1):
      for j in range(1, 6):
          ws6.cell(row=i+1, column=j).value = desc2_subList1[i-1][j-1]  

  #REPORT SUB-DEPT
  ws7 = wb.create_sheet("REPORT-SDEPT")
  report_title_list = ['UPC', 'BRAND', 'DESC1', 'DESC2', 'SIZE', 'FAMILY', 'FNAME', 'REPORT', 'RNAME', 'SDEPT']
  report_subList1 = []
 

  for i in result.keys():
    for j in range(0, len(result[i])):
      if result[i][j][9] != None:
        if 101<=int(result[i][j][9])<=303:
          if result[i][j][7] not in ['30', '93']:
            report_subList1.append(result[i][j])
        elif 401<=int(result[i][j][9])<=905:
          if result[i][j][7] not in ['40', '94']:
            report_subList1.append(result[i][j])
        elif 1001<=int(result[i][j][9])<=1304:
          if result[i][j][7] not in ['50', '95']:
            report_subList1.append(result[i][j])
        elif 1401<=int(result[i][j][9])<=1501:
          if result[i][j][7] not in ['60', '96']:
            report_subList1.append(result[i][j])
        elif 1601<=int(result[i][j][9])<=3506:
          if result[i][j][7] not in ['11', '12', '13', '14', '15', '16', '17','18', '19']:
            report_subList1.append(result[i][j])                              
        elif 3601<=int(result[i][j][9])<=4105:
          if result[i][j][7] not in ['20']:
            report_subList1.append(result[i][j])
        
  
  report_subList1 = [item for item in report_subList1 if str(item[7] or "x") + str(item[8] or "x")not in ['72CIGARETTE', '71PHONE CARD' ,'70ETC1 (CONSIGNMENT)', '83H/W (CONSIGNMENT)']]


  report_temp = len(report_subList1)
  report_subList1.sort(key=lambda x:x[0])

  for i in range(1, 11):
      ws7.cell(row=1, column=i).value = report_title_list[i-1]
      ws7.cell(row=1, column=i).alignment = Alignment(horizontal = 'center', vertical = 'center')
      ws7.cell(row=1, column=i).font = Font(bold=True)

  for i in range(1, report_temp+1):
      for j in range(1, 11):
          ws7.cell(row=i+1, column=j).value = report_subList1[i-1][j-1]  


  #TAX
  tax_title_list = ['UPC', 'BRAND', 'DESC1', 'DESC2', 'SIZE', 'REPORT', 'RNAME', 'SDEPT', 'SNAME','CATEGORY', 'CNAME', 'TAX1' ,'TAX2', 'FSTMP']
  #GA55
  if 'GA55' in f:
    ws8 = wb.create_sheet("TAX1")
    ga55_tax1 = []
    for i in result.keys():
      for j in range(0, len(result[i])):
        if result[i][j][13]!=None and result[i][j][14] == None:
          if result[i][j][9] == None:
            if result[i][j][8] in ['H/W', 'CIGARETTE', 'DELI'] or result[i][j][10] == 'CLEANING AND LAUNDRY':
              ga55_tax1.append(result[i][j])
          elif result[i][j][8] in ['H/W', 'CIGARETTE', 'DELI'] or 3601<=int(result[i][j][9])<=4105 or result[i][j][10] == 'CLEANING AND LAUNDRY':
            ga55_tax1.append(result[i][j])


    ga55_tax1 = [item for item in ga55_tax1 if item[8] != 'DELI']
    ga55_tax1 = [item for item in ga55_tax1 if item[0] not in['0076173473042', '0081244103011', '0081244103012', '0081244103013', '0081244103014', '0081244103015', '0081244103016', '0081244103017', '0081244103018', '0081244103100', '0081244103101', '0085084100731', '0085084100732', '0085084100733', '0085084100734', '0085084100735', '0085084100737', '0085084100752', '0085084100755', '0086084700033', '0094502785125', '0400693700007']]
    ga55_tax1 = [item for item in ga55_tax1 if item[0] not in ['0718939287781', '0880103330852', '0880509796426', '0880509796432', '0880591553023', '0880591553342', '0880591553382', '0880591553422', '0880777909568', '0880905281148', '0880913133864', '0880915561130', '0880916100701', '0880920164325', '0880922198001', '0880922198007', '0880922198011', '0880922198041', '0880922198101', '0880922198222', '0880922198229', '0880925364002', '0880925364636', '0880940335092', '0880940335093', '0880940335194', '0880925364010']]
    ga55_tax1_temp = len(ga55_tax1)
    ga55_tax1.sort(key=lambda x:x[0])

    for i in range(1, 15):
        ws8.cell(row=1, column=i).value = tax_title_list[i-1]
        ws8.cell(row=1, column=i).alignment = Alignment(horizontal = 'center', vertical = 'center')
        ws8.cell(row=1, column=i).font = Font(bold=True)

    for i in range(1, ga55_tax1_temp+1):
        for j in range(1, 6):
            ws8.cell(row=i+1, column=j).value = ga55_tax1[i-1][j-1] 

        for j in range(6, 15):
            ws8.cell(row=i+1, column=j).value = ga55_tax1[i-1][j+1]             


    ws9 = wb.create_sheet("TAX2")
    ga55_tax2 = []
    for i in result.keys():
      for j in range(0, len(result[i])):
        if result[i][j][13]==None and result[i][j][14] != None:
          if result[i][j][8] not in ['H/W', 'CIGARETTE', 'DELI'] and result[i][j][10] not in ['CLEANING AND LAUNDRY', 'CIGARETTES', 'WINE'] and result[i][j][12] not in ['CLEANING AND LAUNDRY','CIGARETTES', 'WINE', 'BEER'] and result[i][j][10] != 'CONSIGNMENT' and (int(result[i][j][9])<3001 or int(result[i][j][9]) > 4005):
            ga55_tax2.append(result[i][j])



    ga55_tax2 = [item for item in ga55_tax2 if item[0] not in ['0005000050389', '0005000057216', '0005000083612', '0020064600000', '0020098100000', '0020164600000', '0020499900000', '0188061281032']]
    ga55_tax2_temp = len(ga55_tax2)
    ga55_tax2.sort(key=lambda x:x[0])

    for i in range(1, 15):
        ws9.cell(row=1, column=i).value = tax_title_list[i-1]
        ws9.cell(row=1, column=i).alignment = Alignment(horizontal = 'center', vertical = 'center')
        ws9.cell(row=1, column=i).font = Font(bold=True)

    for i in range(1, ga55_tax2_temp+1):
        for j in range(1, 6):
            ws9.cell(row=i+1, column=j).value = ga55_tax2[i-1][j-1] 

        for j in range(6, 15):
            ws9.cell(row=i+1, column=j).value = ga55_tax2[i-1][j+1]        
            
    ws10 = wb.create_sheet("TAX1,2")
    ga55_tax12 = []
    for i in result.keys():
      for j in range(0, len(result[i])):
        if result[i][j][13] != None and result[i][j][14] != None:
          ga55_tax12.append(result[i][j])

   
      
    ga55_tax12 = [item for item in ga55_tax12 if item[0] not in ['0064455864957']]
    ga55_tax12_temp = len(ga55_tax12)
    ga55_tax12.sort(key=lambda x:x[0])

    for i in range(1, 15):
        ws10.cell(row=1, column=i).value = tax_title_list[i-1]
        ws10.cell(row=1, column=i).alignment = Alignment(horizontal = 'center', vertical = 'center')
        ws10.cell(row=1, column=i).font = Font(bold=True)

    for i in range(1, ga55_tax12_temp+1):
        for j in range(1, 6):
            ws10.cell(row=i+1, column=j).value = ga55_tax12[i-1][j-1] 

        for j in range(6, 15):
            ws10.cell(row=i+1, column=j).value = ga55_tax12[i-1][j+1]        

  #TAX
  #IL70
  if 'IL70' in f:
    ws8 = wb.create_sheet("TAX1")
    il70_tax1 = []
    for i in result.keys():
      for j in range(0, len(result[i])):
        if result[i][j][13]!=None and result[i][j][14] == None:
          if result[i][j][9] == None:
            if result[i][j][8] == 'H/W':
              il70_tax1.append(result[i][j])
          elif result[i][j][8] == 'H/W' or 3601<=int(result[i][j][9])<=4105:
            il70_tax1.append(result[i][j])


    il70_tax1_temp = len(il70_tax1)
    il70_tax1.sort(key=lambda x:x[0])

    for i in range(1, 15):
        ws8.cell(row=1, column=i).value = tax_title_list[i-1]
        ws8.cell(row=1, column=i).alignment = Alignment(horizontal = 'center', vertical = 'center')
        ws8.cell(row=1, column=i).font = Font(bold=True)

    for i in range(1, il70_tax1_temp+1):
        for j in range(1, 6):
            ws8.cell(row=i+1, column=j).value = il70_tax1[i-1][j-1] 
        for j in range(6, 15):
            ws8.cell(row=i+1, column=j).value = il70_tax1[i-1][j+1] 

    ws9 = wb.create_sheet("TAX2")
    il70_tax2 = []
    for i in result.keys():
      for j in range(0, len(result[i])):
        if result[i][j][13]==None and result[i][j][14] != None:
          if result[i][j][8] in ['MEAT', 'PRODUCE', 'FISH'] and (int(result[i][j][9])<3601 or int(result[i][j][9])>4105):
            if result[i][j][0] not in ['0002874462642', '0002874462674', '0002874472206', '0002874472222', '0002874472223', '0002874472224', '0002874472326',
             '0002874472640', '0002874472824', '0002874472882', '0020099500000' ] :
             il70_tax2.append(result[i][j])
 
    il70_tax2 = [item for item in il70_tax2 if item[0] not in ['0002874472202', '0002874472200', '0002874472214','0020275500000', '0020287300000', '0020288000000', '0020289200000', '0008165294151']]
    il70_tax2_temp = len(il70_tax2)
    il70_tax2.sort(key=lambda x:x[0])

    for i in range(1, 15):
        ws9.cell(row=1, column=i).value = tax_title_list[i-1]
        ws9.cell(row=1, column=i).alignment = Alignment(horizontal = 'center', vertical = 'center')
        ws9.cell(row=1, column=i).font = Font(bold=True)

    for i in range(1, il70_tax2_temp+1):
        for j in range(1, 6):
            ws9.cell(row=i+1, column=j).value = il70_tax2[i-1][j-1]
        for j in range(6, 15):
            ws9.cell(row=i+1, column=j).value = il70_tax2[i-1][j+1]


    ws10 = wb.create_sheet("TAX1,2")
    il70_tax12 = []
    for i in result.keys():
      for j in range(0, len(result[i])):
        if result[i][j][13] != None and result[i][j][14] != None:
          il70_tax12.append(result[i][j])


    il70_tax12_temp = len(il70_tax12)
    il70_tax12.sort(key=lambda x:x[0])

    for i in range(1, 15):
        ws10.cell(row=1, column=i).value = tax_title_list[i-1]
        ws10.cell(row=1, column=i).alignment = Alignment(horizontal = 'center', vertical = 'center')
        ws10.cell(row=1, column=i).font = Font(bold=True)

    for i in range(1, il70_tax12_temp+1):
        for j in range(1, 6):
            ws10.cell(row=i+1, column=j).value = il70_tax12[i-1][j-1]
        for j in range(6, 15):
            ws10.cell(row=i+1, column=j).value = il70_tax12[i-1][j+1]
  #TAX
  #PA88
  if 'PA88' in f:
    ws8 = wb.create_sheet("TAX1")
    pa88_tax1 = []
    for i in result.keys():
      for j in range(0, len(result[i])):
        if result[i][j][13]!=None and result[i][j][14] == None:
          if result[i][j][9] == '2201':
            if ('미네랄 워터' in result[i][j][3] or '미네랄워터' in result[i][j][3]) and '탄산' not in result[i][j][3]:
              pa88_tax1.append(result[i][j])
          elif result[i][j][9] != None:
            if 'H/W' not in result[i][j][8] and 'PHONE CARD' not in result[i][j][8] and ( int(result[i][j][9])<3001 or 3403<int(result[i][j][9])<3601 or int(result[i][j][9]) > 4105) and result[i][j][10] != 'CIGARETTES' and result[i][j][10] != 'WINE':
               pa88_tax1.append(result[i][j])

           

  
    pa88_tax1 = [item for item in pa88_tax1 if item[7] not in ['30', '50', '60']]

    pa88_tax1_temp = len(pa88_tax1)
    pa88_tax1.sort(key=lambda x:x[0])

    for i in range(1, 15):
        ws8.cell(row=1, column=i).value = tax_title_list[i-1]
        ws8.cell(row=1, column=i).alignment = Alignment(horizontal = 'center', vertical = 'center')
        ws8.cell(row=1, column=i).font = Font(bold=True)

    for i in range(1, pa88_tax1_temp+1):
        for j in range(1, 6):
            ws8.cell(row=i+1, column=j).value = pa88_tax1[i-1][j-1] 
        for j in range(6, 15):
            ws8.cell(row=i+1, column=j).value = pa88_tax1[i-1][j+1] 


    ws9 = wb.create_sheet("TAX2")
    pa88_tax2 = []
    for i in result.keys():
      for j in range(0, len(result[i])):
        if result[i][j][13]==None and result[i][j][14] != None:
          pa88_tax2.append(result[i][j])


    pa88_tax2_temp = len(pa88_tax2)
    pa88_tax2.sort(key=lambda x:x[0])

    for i in range(1, 15):
        ws9.cell(row=1, column=i).value = tax_title_list[i-1]
        ws9.cell(row=1, column=i).alignment = Alignment(horizontal = 'center', vertical = 'center')
        ws9.cell(row=1, column=i).font = Font(bold=True)

    for i in range(1, pa88_tax2_temp+1):
        for j in range(1, 6):
            ws9.cell(row=i+1, column=j).value = pa88_tax2[i-1][j-1]
        for j in range(6, 15):
            ws9.cell(row=i+1, column=j).value = pa88_tax2[i-1][j+1]
            
    ws10 = wb.create_sheet("TAX1,2")
    pa88_tax12 = []
    for i in result.keys():
      for j in range(0, len(result[i])):
        if result[i][j][13] != None and result[i][j][14] != None:
          pa88_tax12.append(result[i][j])


    pa88_tax12_temp = len(pa88_tax12)
    pa88_tax12.sort(key=lambda x:x[0])

    for i in range(1, 15):
        ws10.cell(row=1, column=i).value = tax_title_list[i-1]
        ws10.cell(row=1, column=i).alignment = Alignment(horizontal = 'center', vertical = 'center')
        ws10.cell(row=1, column=i).font = Font(bold=True)

    for i in range(1, pa88_tax12_temp+1):
        for j in range(1, 6):
            ws10.cell(row=i+1, column=j).value = pa88_tax12[i-1][j-1]
        for j in range(6, 15):
            ws10.cell(row=i+1, column=j).value = pa88_tax12[i-1][j+1]

    ws11 = wb.create_sheet("FSTMP")
    pa88_fstmp = []
    for i in result.keys():
      for j in range(0, len(result[i])):
        if result[i][j][9] =='2201' and result[i][j][15] == None:
          pa88_fstmp.append(result[i][j])


    pa88_fstmp_temp = len(pa88_fstmp)
    pa88_fstmp.sort(key=lambda x:x[0])

    for i in range(1, 15):
        ws11.cell(row=1, column=i).value = tax_title_list[i-1]
        ws11.cell(row=1, column=i).alignment = Alignment(horizontal = 'center', vertical = 'center')
        ws11.cell(row=1, column=i).font = Font(bold=True)

    for i in range(1, pa88_fstmp_temp+1):
        for j in range(1, 6):
            ws11.cell(row=i+1, column=j).value = pa88_fstmp[i-1][j-1]
        for j in range(6, 15):
            ws11.cell(row=i+1, column=j).value = pa88_fstmp[i-1][j+1]            
  wb.save(f)

